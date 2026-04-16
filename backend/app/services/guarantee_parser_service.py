"""
Guarantee Statement Parser - Uses Claude Vision to extract guarantee items from bank guarantee documents.

Supports:
- PDF guarantee statements (converted to images)
- Excel guarantee reports

Extracts per item:
- Buyer/beneficiary name
- Guarantee type (חוק מכר, ביצוע, כספית, בנקאית)
- Original amount
- Indexed balance
- Expiry/end date
- Apartment number (if applicable)
"""
import os
import base64
import json
import logging
import tempfile
from typing import Dict, List, Any
from datetime import datetime
from decimal import Decimal
import anthropic

logger = logging.getLogger(__name__)

GUARANTEE_TYPE_MAP = {
    'ערבות חוק מכר': 'sale_law',
    'חוק מכר': 'sale_law',
    'ערבות ביצוע': 'performance',
    'ביצוע': 'performance',
    'ערבות כספית': 'financial',
    'כספית': 'financial',
    'ערבות בנקאית': 'bank',
    'בנקאית': 'bank',
}

GUARANTEE_TYPE_LABELS = {
    'sale_law': 'חוק מכר',
    'performance': 'ביצוע',
    'financial': 'כספית',
    'bank': 'בנקאית',
    'other': 'אחר',
}


class GuaranteeParserService:
    """AI-powered guarantee document parser using Claude Vision."""

    def __init__(self):
        self._client = None

    def _get_client(self) -> anthropic.Anthropic:
        if self._client is None:
            api_key = os.environ.get('ANTHROPIC_API_KEY', '')
            if not api_key:
                raise ValueError("ANTHROPIC_API_KEY לא מוגדר")
            self._client = anthropic.Anthropic(api_key=api_key)
        return self._client

    def parse_guarantee_file(
        self, file_content: bytes, file_type: str, filename: str,
    ) -> Dict[str, Any]:
        """
        Parse a guarantee statement and return structured items.

        Returns:
            {
                "items": [
                    {
                        "buyer_name": "ישראל ישראלי",
                        "guarantee_type": "sale_law",
                        "original_amount": 500000,
                        "indexed_balance": 520000,
                        "expiry_date": "2026-12-31",
                        "apartment_number": "12",
                        "notes": ""
                    }, ...
                ],
                "total_balance": 2600000,
                "warnings": []
            }
        """
        if file_type in [
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ]:
            return self._parse_excel(file_content, filename)
        elif file_type == 'application/pdf':
            return self._parse_pdf(file_content, filename)
        else:
            raise ValueError(f"סוג קובץ לא נתמך: {file_type}")

    def _parse_excel(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse Excel guarantee report."""
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        # Try to find header row by scanning first 10 rows
        header_row = None
        header_map: Dict[str, int] = {}
        keywords = ['מוטב', 'שם', 'סוג', 'סכום', 'יתרה', 'תאריך', 'דירה', 'ערבות']

        for row_idx in range(1, min(ws.max_row + 1, 15)):
            cells = [str(ws.cell(row=row_idx, column=c).value or '').strip() for c in range(1, ws.max_column + 1)]
            matches = sum(1 for cell in cells for kw in keywords if kw in cell)
            if matches >= 2:
                header_row = row_idx
                for col_idx, cell_val in enumerate(cells):
                    header_map[cell_val.lower()] = col_idx
                break

        if header_row is None:
            # Fall back to AI parsing
            return self._parse_with_ai(file_content, filename, 'excel')

        # Map columns
        def find_col(*terms: str):
            for h, idx in header_map.items():
                for t in terms:
                    if t in h:
                        return idx
            return None

        buyer_col = find_col('מוטב', 'שם', 'רוכש', 'לקוח')
        type_col = find_col('סוג', 'ערבות')
        amount_col = find_col('סכום', 'קרן', 'מקור')
        balance_col = find_col('יתרה', 'צמוד')
        date_col = find_col('תוקף', 'תאריך סיום', 'פקיעה')
        apt_col = find_col('דירה', 'יחידה', 'נכס')

        items = []
        warnings = []
        total = Decimal('0')

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row):
                continue

            buyer = str(row[buyer_col]).strip() if buyer_col is not None and row[buyer_col] else ''
            if not buyer or buyer == 'None':
                continue

            raw_type = str(row[type_col]).strip() if type_col is not None and row[type_col] else ''
            g_type = 'other'
            for key, val in GUARANTEE_TYPE_MAP.items():
                if key in raw_type:
                    g_type = val
                    break

            orig = self._to_decimal(row[amount_col] if amount_col is not None else None)
            bal = self._to_decimal(row[balance_col] if balance_col is not None else None) or orig
            total += bal

            expiry = None
            if date_col is not None and row[date_col]:
                expiry = self._parse_date(row[date_col])

            apt_num = str(row[apt_col]).strip() if apt_col is not None and row[apt_col] else ''
            if apt_num == 'None':
                apt_num = ''

            items.append({
                'buyer_name': buyer,
                'guarantee_type': g_type,
                'original_amount': float(orig),
                'indexed_balance': float(bal),
                'expiry_date': expiry,
                'apartment_number': apt_num,
                'notes': '',
            })

        return {'items': items, 'total_balance': float(total), 'warnings': warnings}

    def _parse_pdf(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Parse PDF guarantee — text extraction first, AI vision fallback."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise ValueError("PyMuPDF is not installed. pip install pymupdf")

        doc = fitz.open(stream=file_content, filetype="pdf")
        full_text = ""
        for page in doc:
            full_text += page.get_text() + "\n"
        doc.close()

        # If PDF has extractable text, parse deterministically
        if len(full_text.strip()) > 50:
            try:
                result = self._parse_guarantee_pdf_text(full_text)
                if result and result.get('items'):
                    logger.info(f"Guarantee PDF text parser extracted {len(result['items'])} items")
                    return result
            except Exception as e:
                logger.warning(f"Guarantee PDF text parser failed: {e}")

        # Fallback to AI
        return self._parse_with_ai(file_content, filename, 'pdf')

    def _parse_guarantee_pdf_text(self, text: str) -> Dict[str, Any]:
        """Deterministic parser for guarantee PDFs with extractable text.

        Guarantee statements typically list items in a table with columns:
        buyer name, guarantee type, amount, indexed balance, expiry date, apartment number.
        """
        import re

        items = []
        warnings = []
        total = Decimal('0')

        # Strategy: find lines with monetary amounts and dates that look like guarantee items.
        # Each guarantee item usually has: a name, a monetary amount, and a date.

        lines = text.split('\n')
        lines = [l.strip() for l in lines if l.strip()]

        # Detect guarantee type context from text
        def detect_type_from_context(nearby_text: str) -> str:
            nearby_lower = nearby_text.lower()
            for key, val in GUARANTEE_TYPE_MAP.items():
                if key in nearby_text:
                    return val
            return 'sale_law'  # Default for bank monitoring reports

        # Pattern 1: Tabular data — look for rows with amounts and dates
        # Common Israeli guarantee statement format:
        # buyer_name | apt_num | amount | indexed_amount | expiry_date
        amount_pattern = re.compile(r'([\d,]+\.?\d{0,2})')
        date_pattern = re.compile(r'(\d{2}[./]\d{2}[./]\d{4})')

        # Try to find a header row to understand structure
        header_idx = None
        for i, line in enumerate(lines):
            # Header indicators
            if any(kw in line for kw in ['מוטב', 'שם הרוכש', 'שם הקונה']) and \
               any(kw in line for kw in ['סכום', 'יתרה', 'ערבות']):
                header_idx = i
                break

        # If we found a header, parse subsequent rows
        if header_idx is not None:
            for i in range(header_idx + 1, len(lines)):
                line = lines[i]
                # Skip empty-looking lines, footers, etc.
                if len(line) < 5:
                    continue
                if any(kw in line for kw in ['סה"כ', 'סך הכל', 'עמוד', 'הערות', 'page']):
                    continue

                # Extract amounts from line
                amounts = amount_pattern.findall(line)
                amounts = [float(a.replace(',', '')) for a in amounts if float(a.replace(',', '')) > 100]

                dates = date_pattern.findall(line)

                if not amounts:
                    continue

                # Extract name: text before the first number
                first_num_pos = re.search(r'\d', line)
                name = line[:first_num_pos.start()].strip() if first_num_pos else ''

                # Extract apartment number: small number (1-999)
                apt_num = ''
                small_nums = re.findall(r'\b(\d{1,3})\b', line)
                for sn in small_nums:
                    if 1 <= int(sn) <= 999 and float(sn) not in amounts:
                        apt_num = sn
                        break

                g_type = detect_type_from_context(text[:500])
                orig = amounts[0] if len(amounts) >= 1 else 0
                bal = amounts[1] if len(amounts) >= 2 else orig
                total += Decimal(str(bal))

                expiry = None
                if dates:
                    expiry = self._parse_date(dates[0])

                items.append({
                    'buyer_name': name,
                    'guarantee_type': g_type,
                    'original_amount': orig,
                    'indexed_balance': bal,
                    'expiry_date': expiry,
                    'apartment_number': apt_num,
                    'notes': '',
                })
        else:
            # No header found — try a more aggressive scan
            # Look for lines with Hebrew names + amounts
            for i, line in enumerate(lines):
                if len(line) < 10:
                    continue

                amounts = amount_pattern.findall(line)
                amounts = [float(a.replace(',', '')) for a in amounts if float(a.replace(',', '')) > 1000]
                if not amounts:
                    continue

                # Check if line has Hebrew text (likely a name)
                has_hebrew = bool(re.search(r'[\u0590-\u05FF]', line))
                if not has_hebrew:
                    continue

                first_num_pos = re.search(r'\d', line)
                name = line[:first_num_pos.start()].strip() if first_num_pos else ''
                if not name or len(name) < 3:
                    continue

                dates = date_pattern.findall(line)
                g_type = detect_type_from_context(text[:500])
                orig = amounts[0]
                bal = amounts[1] if len(amounts) >= 2 else orig
                total += Decimal(str(bal))

                items.append({
                    'buyer_name': name,
                    'guarantee_type': g_type,
                    'original_amount': orig,
                    'indexed_balance': bal,
                    'expiry_date': self._parse_date(dates[0]) if dates else None,
                    'apartment_number': '',
                    'notes': '',
                })

        if not items:
            warnings.append('לא נמצאו ערבויות בטקסט המסמך')

        return {
            'items': items,
            'total_balance': float(total),
            'warnings': warnings or ['Parsed from PDF text (no AI)'],
        }

    def _parse_with_ai(self, file_content: bytes, filename: str, source: str) -> Dict[str, Any]:
        """Use Claude Vision to parse guarantee document."""
        client = self._get_client()

        if source == 'pdf':
            content_block = {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": base64.standard_b64encode(file_content).decode("utf-8"),
                },
            }
        else:
            # For Excel, convert to text representation
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            rows_text = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True):
                rows_text.append(' | '.join(str(c or '') for c in row))
            content_block = {
                "type": "text",
                "text": f"Excel file contents:\n" + '\n'.join(rows_text),
            }

        prompt = """אתה מנתח תדפיס ערבויות בנקאיות של פרויקט נדל"ן.

חלץ את כל הערבויות מהמסמך בפורמט JSON הבא:
{
    "items": [
        {
            "buyer_name": "שם המוטב/רוכש",
            "guarantee_type": "sale_law|performance|financial|bank|other",
            "original_amount": 500000,
            "indexed_balance": 520000,
            "expiry_date": "2026-12-31",
            "apartment_number": "12",
            "notes": ""
        }
    ],
    "total_balance": 2600000,
    "warnings": ["אזהרה כלשהי"]
}

סוגי ערבויות:
- sale_law = ערבות חוק מכר
- performance = ערבות ביצוע
- financial = ערבות כספית
- bank = ערבות בנקאית
- other = אחר

החזר JSON בלבד, ללא טקסט נוסף."""

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [content_block, {"type": "text", "text": prompt}],
            }],
        )

        raw = response.content[0].text.strip()
        # Strip markdown code fences if present
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1]
            if raw.endswith('```'):
                raw = raw.rsplit('```', 1)[0]

        try:
            result = json.loads(raw)
            return {
                'items': result.get('items', []),
                'total_balance': result.get('total_balance', 0),
                'warnings': result.get('warnings', []),
            }
        except json.JSONDecodeError:
            logger.error(f"Failed to parse AI response: {raw[:200]}")
            return {'items': [], 'total_balance': 0, 'warnings': ['שגיאה בפענוח תגובת AI']}

    @staticmethod
    def _to_decimal(val) -> Decimal:
        if val is None:
            return Decimal('0')
        try:
            cleaned = str(val).replace(',', '').replace('₪', '').strip()
            return Decimal(cleaned)
        except Exception:
            return Decimal('0')

    @staticmethod
    def _parse_date(val) -> str | None:
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return None


guarantee_parser = GuaranteeParserService()
