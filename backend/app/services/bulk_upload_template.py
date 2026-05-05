"""Generates the bulk-upload Excel template (8 sheets) on demand.

The parser in ``bulk_upload_service.py`` reads cells by position, not by
header text. This generator produces a workbook whose column order
matches that parser exactly, plus Hebrew header labels, example rows,
and inline cell comments listing the allowed enum values.
"""

import io
from datetime import date

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
_EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_INSTR_FILL = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
_THIN = Side(border_style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _add_data_sheet(wb, title, headers, example_row, comments=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 4)
        if comments and col_idx - 1 < len(comments) and comments[col_idx - 1]:
            c.comment = Comment(comments[col_idx - 1], "Maakav")
    for col_idx, v in enumerate(example_row, 1):
        c = ws.cell(row=2, column=col_idx, value=v)
        c.fill = _EXAMPLE_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = _BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def build_template_bytes() -> bytes:
    """Build the workbook in memory and return its bytes."""
    wb = openpyxl.Workbook()

    # Sheet 0 — instructions
    ws = wb.active
    ws.title = "הוראות"
    ws.sheet_view.rightToLeft = True
    instructions = [
        ("תבנית העלאת פרויקט חדש - מערכת מעקב", True),
        ("", False),
        ("הוראות שימוש:", True),
        ("1. כל גיליון מייצג מקור נתונים שונה בפרויקט.", False),
        ("2. שורה 1 בכל גיליון = כותרות (אין לשנות סדר עמודות).", False),
        ("3. שורה 2 ואילך = הנתונים שלך.", False),
        ("4. ניתן להשאיר שדות אופציונליים ריקים.", False),
        ("5. שדות מוארים בצהוב = דוגמה (יש למחוק לפני העלאה).", False),
        ("", False),
        ("גיליונות הקובץ:", True),
        ("• מלאי מגורים יזם - דירות שבבעלות היזם (להשכרה/מכירה)", False),
        ("• מלאי מגורים בעלים - דירות שבבעלות דיירים (תמורה)", False),
        ("• מלאי שאינו מגורים יזם - מסחר/משרדים בבעלות היזם", False),
        ("• מלאי שאינו מגורים בעלים - מסחר/משרדים בבעלות דיירים", False),
        ("• תקציב הפרויקט - קטגוריות תקציב וסעיפים", False),
        ("• ערבויות - ערבויות בנקאיות לרוכשים", False),
        ("• חשבון ליווי סגור - פרטי חשבון ליווי, הלוואות ופקדונות", False),
        ("", False),
        ("טיפ: ניתן למחוק גיליונות שאינם רלוונטיים לפרויקט.", True),
        ("העלאה: כניסה לפרויקט ← העלה Excel", False),
    ]
    for i, (text, bold) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if bold else 11, name="Arial")
        cell.alignment = Alignment(horizontal="right", wrap_text=True)
        if bold and text:
            cell.fill = _INSTR_FILL
    ws.column_dimensions["A"].width = 80

    # Sheet 1 — Developer residential
    _add_data_sheet(
        wb,
        "מלאי מגורים יזם",
        ["בניין", "אגף", "קומה", "מס׳ יחידה", "מס׳ תכנית", "כיוון",
         "סוג יחידה", "חדרים", "שטח נטו (מ״ר)", "מרפסת (מ״ר)", "גינה/גג (מ״ר)",
         "חניות", "מחסנים", "מחיר/מ״ר משוקלל", "מחיר/מ״ר",
         "מחיר כולל מע״מ", "מחיר ללא מע״מ", "סטטוס"],
        ["A", "צפון", "3", "12", "P3-12", "צפון-מזרח",
         "דירה", 4, 95.5, 12.0, 0,
         1, 1, 28000, 30000,
         2865000, 2448718, "לשיווק"],
        comments=[
            "שם/מספר בניין (חובה)", None, None, None, None, None,
            "ערכים מותרים: דירה, פנטהאוז, גן, דופלקס, דופלקס גן, דופלקס גג, מיני פנטהאוז",
            None, None, None, None, None, None, None, None, None, None,
            "ערכים מותרים: לשיווק, נמכר, שמור, להשכרה, תמורה",
        ],
    )

    # Sheet 2 — Resident residential
    _add_data_sheet(
        wb,
        "מלאי מגורים בעלים",
        ["בניין", "אגף", "קומה", "מס׳ יחידה", "מס׳ תכנית", "כיוון",
         "סוג יחידה", "חדרים", "שטח נטו (מ״ר)", "מרפסת (מ״ר)", "גינה/גג (מ״ר)",
         "חניות", "מחסנים", "מחיר/מ״ר משוקלל", "מחיר/מ״ר",
         "מחיר כולל מע״מ", "שם בעלים"],
        ["A", "דרום", "1", "5", "P1-05", "מערב",
         "דירה", 3, 80.0, 8.0, 0,
         1, 1, 25000, 27000,
         2160000, "ישראל ישראלי"],
        comments=[None] * 16 + ["שם הבעלים הקיים (תמורה)"],
    )

    # Sheet 3 — Developer commercial
    _add_data_sheet(
        wb,
        "מלאי שאינו מגורים יזם",
        ["בניין", "אגף", "קומה", "מס׳ יחידה", "מס׳ תכנית", "כיוון",
         "סוג יחידה", "תת-סוג", "חניות", "שטח נטו (מ״ר)", "שטח ברוטו (מ״ר)",
         "מרפסת (מ״ר)", "גלריה (מ״ר)", "חצר (מ״ר)",
         "מחיר/מ״ר ללא מע״מ", "מחיר/מ״ר עם מע״מ",
         "מחיר כולל מע״מ", "מחיר ללא מע״מ", "סטטוס"],
        ["B", "קרקע", "0", "M-01", "C-01", "צפון",
         "מסחר", "סופרמרקט", 5, 200.0, 220.0,
         0, 0, 0,
         20000, 23400,
         5148000, 4400000, "לשיווק"],
        comments=[None] * 6 + [
            "ערכים מותרים: מסחר, משרדים, מחסן, חניה, תעשיה, לוגיסטיקה, אחר",
        ] + [None] * 11 + ["ערכים: לשיווק, נמכר, שמור, להשכרה, תמורה"],
    )

    # Sheet 4 — Resident commercial
    _add_data_sheet(
        wb,
        "מלאי שאינו מגורים בעלים",
        ["בניין", "אגף", "קומה", "מס׳ יחידה", "מס׳ תכנית", "כיוון",
         "סוג יחידה", "תת-סוג", "חניות", "שטח נטו (מ״ר)", "שטח ברוטו (מ״ר)",
         "מרפסת (מ״ר)", "גלריה (מ״ר)", "חצר (מ״ר)",
         "מחיר/מ״ר ללא מע״מ", "מחיר/מ״ר עם מע״מ",
         "מחיר כולל מע״מ", "מחיר ללא מע״מ", "שם בעלים"],
        ["B", "קרקע", "-1", "S-01", "ST-01", "מערב",
         "מחסן", "", 0, 30.0, 32.0,
         0, 0, 0,
         0, 0, 0, 0, "כהן משה"],
    )

    # Sheet 5 — Budget. Column 3 = pre-project equity investment, recognized
    # by the appraiser. Sum across rows rolls up into the financing tab as
    # "השקעות טרום פרויקט". Leave blank for lines with no pre-spend.
    ws = wb.create_sheet("תקציב הפרויקט")
    ws.sheet_view.rightToLeft = True
    headers = ["קטגוריה", "סעיף", "השקעות הון עצמי", 'עלות ללא מע"מ בש"ח', "הערה"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=1, column=1).comment = Comment(
        "ערכים מותרים: הוצאות דיירים, קרקע ומיסוי, הוצאות עקיפות, בניה ישירה",
        "Maakav",
    )
    ws.cell(row=1, column=3).comment = Comment(
        "סכום הון עצמי שכבר הושקע בפרויקט עבור הסעיף, המוכר על ידי משרד שמאים. ניתן להשאיר ריק.",
        "Maakav",
    )
    examples = [
        ["הוצאות דיירים", "מארגן דיירים", 20000, 250000, ""],
        ["קרקע ומיסוי", "תיווך", 100000, 1250000, ""],
        ["קרקע ומיסוי", "רכישת פרויקט", 3000000, 30000000, ""],
        ["הוצאות עקיפות", "תכנון ויועצים", 23756, 300000, ""],
        ["בניה ישירה", "בנייה כללית", "", 18500000, "חוזה ראשי"],
        ["בניה ישירה", "מערכות חשמל", "", 950000, ""],
    ]
    for r_idx, row in enumerate(examples, 2):
        for c_idx, v in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            c.fill = _EXAMPLE_FILL
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Sheet 6 — Guarantees (special: total at A4, headers at row 10, data row 11+)
    ws = wb.create_sheet("ערבויות")
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value="ערבויות חוק מכר - הנפקות לרוכשים").font = Font(bold=True, size=14, name="Arial")
    ws.cell(row=3, column=1, value="סך יתרה צמודה כוללת:").font = Font(bold=True, name="Arial")
    cell = ws.cell(row=4, column=1, value=4500000)
    cell.fill = _EXAMPLE_FILL
    cell.comment = Comment("שורה זו (תא A4) חייבת להכיל את סך כל הערבויות הצמוד", "Maakav")
    g_headers = ["סוג ערבות", "מס׳ דירה", "שם רוכש", "סכום מקורי",
                 "תאריך תחילת חישוב", "בסיס הצמדה", "סוג הצמדה",
                 "תאריך פקיעה", "יתרה צמודה", "—", "הערות"]
    for i, h in enumerate(g_headers, 1):
        c = ws.cell(row=10, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.cell(row=10, column=1).comment = Comment(
        "ערכים: ערבות ביצוע, ערבות חוק מכר, ערבות כספית, ערבות בנקאית", "Maakav",
    )
    g_examples = [
        ["ערבות חוק מכר", "12", "כהן דוד", 1200000, "2024-03-15", "מדד תשומות הבנייה", "מדד", "2027-03-15", 1280000, "", "רוכש דירה"],
        ["ערבות חוק מכר", "15", "לוי שרה", 1500000, "2024-05-01", "מדד המחירים לצרכן", "מדד", "2027-05-01", 1560000, "", ""],
        ["ערבות ביצוע", "", "בנק לאומי", 1800000, "2024-01-01", "ללא הצמדה", "ללא", "2026-12-31", 1660000, "", "ערבות ביצוע פרויקט"],
    ]
    for r_idx, row in enumerate(g_examples, 11):
        for c_idx, v in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            c.fill = _EXAMPLE_FILL
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center")
    for col_idx in range(1, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["F"].width = 22

    # Sheet 7 — Financing (חשבון ליווי סגור) — fixed positional layout
    ws = wb.create_sheet("חשבון ליווי סגור")
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value="חשבון ליווי - פרטי חשבון, יתרת פתיחה, הלוואות ופקדונות").font = Font(bold=True, size=14, name="Arial")

    def _label(r, c, text):
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="right")

    def _example(r, c, val):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = _EXAMPLE_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER

    _label(2, 1, "מס׳ חשבון:");           _example(2, 2, "767-220100/80")
    _label(3, 1, "קוד בנק:");              _example(3, 2, "10")
    _label(4, 1, "סניף:");                  _example(4, 2, "800")
    _label(6, 1, "יתרת פתיחה")
    _label(6, 2, "תאריך"); _label(6, 3, "זכות"); _label(6, 4, "חובה")
    _example(7, 2, date(2024, 1, 1)); _example(7, 3, 5000000); _example(7, 4, 0)
    _label(10, 1, "הלוואות בכירות")
    _label(10, 2, "תאריך"); _label(10, 3, "קרן"); _label(10, 4, "יתרה")
    _example(11, 2, date(2024, 1, 15)); _example(11, 3, 20000000); _example(11, 4, 19500000)
    _example(12, 2, date(2024, 6, 1));  _example(12, 3, 5000000);  _example(12, 4, 4900000)
    _label(16, 1, "הלוואות נדחות")
    _label(16, 2, "תאריך"); _label(16, 3, "קרן"); _label(16, 4, "יתרה")
    _example(17, 2, date(2024, 2, 10)); _example(17, 3, 3000000); _example(17, 4, 3000000)
    _label(21, 1, "פקדונות (פק״מ)")
    _label(21, 2, "תאריך"); _label(21, 3, "קרן"); _label(21, 4, "יתרה")
    _example(22, 2, date(2024, 4, 1)); _example(22, 3, 1500000); _example(22, 4, 1518000)

    ws.column_dimensions["A"].width = 22
    for col in "BCD":
        ws.column_dimensions[col].width = 18
    ws.cell(row=1, column=1).comment = Comment(
        "מבנה הגיליון קבוע לפי מיקום שורה+עמודה. שמרו על מיקומי שורות 2-4 (פרטי חשבון), 7 (פתיחה), 11-13 (בכירות), 17-18 (נדחות), 22-23 (פקדונות).",
        "Maakav",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
