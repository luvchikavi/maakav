"""
Bulk Upload Service - parses the unified Excel workbook (8 tabs)
and returns a structured preview for user confirmation before saving.

Tabs:
1. חשבון ליווי סגור → ProjectFinancing (account info + loan balances)
2. ערבויות → GuaranteeSnapshot items
3. פרוגרמה → skipped (not used in calculations)
4. מלאי מגורים יזם → Apartment (developer, residential)
5. מלאי שאינו מגורים יזם → Apartment (developer, non-residential)
6. מלאי מגורים בעלים → Apartment (resident, residential)
7. מלאי שאינו מגורים בעלים → Apartment (resident, non-residential)
8. תקציב הפרויקט → BudgetCategory + BudgetLineItem
"""

import io
import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# ── Hebrew mappings ──────────────────────────────────────────

UNIT_TYPE_MAP = {
    "דירה": "apartment", "פנטהאוז": "penthouse", "גן": "garden",
    "דופלקס": "duplex", "דופלקס גן": "duplex_garden", "דופלקס גג": "duplex_roof",
    "מיני פנטהאוז": "mini_penthouse", "משרדים": "office", "מסחר": "retail",
    "מחסן": "storage", "חניה": "parking", "תעשיה": "other",
    "לוגיסטיקה": "other", "אחר": "other",
}

STATUS_MAP = {
    "לשיווק": "for_sale", "להשכרה": "for_rent", "תמורה": "compensation",
    "שמור": "reserved", "נמכר": "sold",
}

CATEGORY_MAP = {
    "הוצאות דיירים": "tenant_expenses",
    "קרקע ומיסוי": "land_and_taxes",
    "הוצאות עקיפות": "indirect_costs",
    "בניה ישירה": "direct_construction",
    "בנייה ישירה": "direct_construction",
}

GUARANTEE_TYPE_MAP = {
    "ערבות ביצוע": "performance", "ערבות חוק מכר": "sale_law",
    "ערבות כספית": "financial", "ערבות בנקאית": "bank",
}


def _dec(val) -> Decimal | None:
    if val is None:
        return None
    try:
        cleaned = str(val).replace(",", "").replace("₪", "").strip()
        if not cleaned or cleaned == "None":
            return None
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return 0


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ── Main parser ──────────────────────────────────────────────


def parse_bulk_upload(file_content: bytes) -> dict:
    """
    Parse the unified Excel file and return a structured preview.

    Returns:
        {
            "apartments_developer_residential": [...],
            "apartments_developer_commercial": [...],
            "apartments_resident_residential": [...],
            "apartments_resident_commercial": [...],
            "budget_categories": { "tenant_expenses": [...], ... },
            "guarantees": [...],
            "financing": { "account_number": ..., ... },
            "summary": {
                "developer_residential_count": N,
                "developer_commercial_count": N,
                "resident_residential_count": N,
                "resident_commercial_count": N,
                "total_apartments": N,
                "budget_line_count": N,
                "budget_total": N,
                "guarantee_count": N,
                "guarantee_total": N,
                "has_financing": bool,
            },
            "warnings": [...],
            "tabs_found": [...],
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    tabs_found = wb.sheetnames
    warnings = []

    result: dict[str, Any] = {
        "apartments_developer_residential": [],
        "apartments_developer_commercial": [],
        "apartments_resident_residential": [],
        "apartments_resident_commercial": [],
        "budget_categories": {},
        "guarantees": [],
        "financing": None,
        "programme": None,
        "warnings": warnings,
        "tabs_found": tabs_found,
    }

    # ── 1. Residential apartments ────────────────────────────

    for tab_name, ownership, key in [
        ("מלאי מגורים יזם", "developer", "apartments_developer_residential"),
        ("מלאי מגורים בעלים", "resident", "apartments_resident_residential"),
    ]:
        if tab_name not in wb.sheetnames:
            warnings.append(f"טאב '{tab_name}' לא נמצא")
            continue
        ws = wb[tab_name]
        apts = []
        for row in range(2, ws.max_row + 1):
            # A row is valid if any identifying column has data — building may
            # be empty for single-building projects.
            building = _str(ws.cell(row=row, column=1).value)
            wing = _str(ws.cell(row=row, column=2).value)
            floor = _str(ws.cell(row=row, column=3).value)
            unit_number = _str(ws.cell(row=row, column=4).value)
            plan_number = _str(ws.cell(row=row, column=5).value)
            if not any([building, wing, floor, unit_number, plan_number]):
                continue

            last_col = ws.max_column
            last_val = _str(ws.cell(row=row, column=last_col).value)

            apt = {
                "building_number": building or "A",
                "wing": _str(ws.cell(row=row, column=2).value),
                "floor": _str(ws.cell(row=row, column=3).value),
                "unit_number": _str(ws.cell(row=row, column=4).value),
                "plan_number": _str(ws.cell(row=row, column=5).value),
                "direction": _str(ws.cell(row=row, column=6).value),
                "unit_type": UNIT_TYPE_MAP.get(_str(ws.cell(row=row, column=7).value).lower().strip(),
                                                UNIT_TYPE_MAP.get(_str(ws.cell(row=row, column=7).value), "apartment")),
                "room_count": _dec(ws.cell(row=row, column=8).value),
                "net_area_sqm": _dec(ws.cell(row=row, column=9).value),
                "balcony_area_sqm": _dec(ws.cell(row=row, column=10).value),
                "terrace_area_sqm": _dec(ws.cell(row=row, column=11).value),
                "parking_count": _int(ws.cell(row=row, column=12).value),
                "storage_count": _int(ws.cell(row=row, column=13).value),
                "price_per_sqm_equiv": _dec(ws.cell(row=row, column=14).value),
                "price_per_sqm": _dec(ws.cell(row=row, column=15).value),
                "list_price_with_vat": _dec(ws.cell(row=row, column=16).value),
                "ownership": ownership,
                "category": "residential",
            }

            if ownership == "developer":
                apt["list_price_no_vat"] = _dec(ws.cell(row=row, column=17).value)
                apt["unit_status"] = STATUS_MAP.get(last_val.strip(), "for_sale") if last_val else "for_sale"
                apt["owner_name"] = None
            else:
                # Resident: last col = owner name, no price_no_vat column
                apt["list_price_no_vat"] = None
                apt["unit_status"] = "compensation"
                apt["owner_name"] = last_val if last_val else None

            apts.append(apt)
        result[key] = apts

    # ── 2. Non-residential (commercial) apartments ───────────

    for tab_name, ownership, key in [
        ("מלאי שאינו מגורים יזם", "developer", "apartments_developer_commercial"),
        ("מלאי שאינו מגורים בעלים", "resident", "apartments_resident_commercial"),
    ]:
        if tab_name not in wb.sheetnames:
            warnings.append(f"טאב '{tab_name}' לא נמצא")
            continue
        ws = wb[tab_name]
        apts = []
        for row in range(2, ws.max_row + 1):
            building = _str(ws.cell(row=row, column=1).value)
            wing = _str(ws.cell(row=row, column=2).value)
            floor = _str(ws.cell(row=row, column=3).value)
            unit_number = _str(ws.cell(row=row, column=4).value)
            plan_number = _str(ws.cell(row=row, column=5).value)
            if not any([building, wing, floor, unit_number, plan_number]):
                continue

            raw_type = _str(ws.cell(row=row, column=7).value)
            last_col = ws.max_column
            last_val = _str(ws.cell(row=row, column=last_col).value)

            apt = {
                "building_number": building or "A",
                "wing": _str(ws.cell(row=row, column=2).value),
                "floor": _str(ws.cell(row=row, column=3).value),
                "unit_number": _str(ws.cell(row=row, column=4).value),
                "plan_number": _str(ws.cell(row=row, column=5).value),
                "direction": _str(ws.cell(row=row, column=6).value),
                "unit_type": UNIT_TYPE_MAP.get(raw_type.lower().strip(),
                                                UNIT_TYPE_MAP.get(raw_type, "other")),
                "secondary_type": _str(ws.cell(row=row, column=8).value),
                "parking_count": _int(ws.cell(row=row, column=9).value),
                "net_area_sqm": _dec(ws.cell(row=row, column=10).value),
                "gross_area_sqm": _dec(ws.cell(row=row, column=11).value),
                "balcony_area_sqm": _dec(ws.cell(row=row, column=12).value),
                "gallery_area_sqm": _dec(ws.cell(row=row, column=13).value),
                "yard_area_sqm": _dec(ws.cell(row=row, column=14).value),
                "price_per_sqm_no_vat": _dec(ws.cell(row=row, column=15).value),
                "price_per_sqm_with_vat": _dec(ws.cell(row=row, column=16).value),
                "list_price_with_vat": _dec(ws.cell(row=row, column=17).value),
                "list_price_no_vat": _dec(ws.cell(row=row, column=18).value),
                "ownership": ownership,
                "category": "non_residential",
            }

            if ownership == "developer":
                apt["unit_status"] = STATUS_MAP.get(last_val.strip(), "for_sale") if last_val else "for_sale"
                apt["owner_name"] = None
            else:
                apt["unit_status"] = "compensation"
                apt["owner_name"] = last_val if last_val else None

            apts.append(apt)
        result[key] = apts

    # ── 3. Budget ────────────────────────────────────────────

    budget_tab = "תקציב הפרויקט"
    if budget_tab in wb.sheetnames:
        ws = wb[budget_tab]
        categories: dict[str, list] = {}
        for row in range(2, ws.max_row + 1):
            cat_heb = _str(ws.cell(row=row, column=1).value)
            if not cat_heb:
                continue
            cat_key = CATEGORY_MAP.get(cat_heb.strip())
            if not cat_key:
                # Try partial match
                for k, v in CATEGORY_MAP.items():
                    if k in cat_heb:
                        cat_key = v
                        break
            if not cat_key:
                warnings.append(f"קטגוריה לא מוכרת: '{cat_heb}'")
                cat_key = "indirect_costs"

            item = {
                "description": _str(ws.cell(row=row, column=2).value),
                "supplier_name": _str(ws.cell(row=row, column=3).value) or None,
                "cost_no_vat": _dec(ws.cell(row=row, column=4).value),
                "notes": _str(ws.cell(row=row, column=5).value) or None,
            }
            if not item["description"]:
                continue
            categories.setdefault(cat_key, []).append(item)
        result["budget_categories"] = categories
    else:
        warnings.append(f"טאב '{budget_tab}' לא נמצא")

    # ── 4. Guarantees ────────────────────────────────────────

    guarantee_tab = "ערבויות"
    if guarantee_tab in wb.sheetnames:
        ws = wb[guarantee_tab]
        # Find header row (row 10 based on format)
        header_row = None
        for r in range(1, min(ws.max_row + 1, 15)):
            val = _str(ws.cell(row=r, column=1).value)
            if "סוג" in val and "ערבות" in val:
                header_row = r
                break
        if header_row is None:
            header_row = 10  # default

        # Total balance from row 4
        total_balance = _dec(ws.cell(row=4, column=1).value)

        items = []
        for row in range(header_row + 1, ws.max_row + 1):
            g_type_heb = _str(ws.cell(row=row, column=1).value)
            if not g_type_heb:
                continue
            g_type = "other"
            for k, v in GUARANTEE_TYPE_MAP.items():
                if k in g_type_heb:
                    g_type = v
                    break

            item = {
                "guarantee_type": g_type,
                "apartment_number": _str(ws.cell(row=row, column=2).value),
                "buyer_name": _str(ws.cell(row=row, column=3).value),
                "original_amount": float(_dec(ws.cell(row=row, column=4).value) or 0),
                "calculation_start_date": _date(ws.cell(row=row, column=5).value),
                "indexation_base": _str(ws.cell(row=row, column=6).value),
                "indexation_type": _str(ws.cell(row=row, column=7).value),
                "expiry_date": _date(ws.cell(row=row, column=8).value),
                "indexed_balance": float(_dec(ws.cell(row=row, column=9).value) or 0),
                "notes": _str(ws.cell(row=row, column=11).value) or "",
            }
            items.append(item)

        result["guarantees"] = items
    else:
        warnings.append(f"טאב '{guarantee_tab}' לא נמצא")

    # ── 5. Financing (bank account) ──────────────────────────

    financing_tab = "חשבון ליווי סגור"
    if financing_tab in wb.sheetnames:
        ws = wb[financing_tab]
        result["financing"] = {
            "account_number": _str(ws.cell(row=2, column=2).value),
            "bank_code": _str(ws.cell(row=3, column=2).value),
            "branch": _str(ws.cell(row=4, column=2).value),
            "opening_balance_date": _date(ws.cell(row=7, column=2).value),
            "opening_balance_credit": _dec(ws.cell(row=7, column=3).value),
            "opening_balance_debit": _dec(ws.cell(row=7, column=4).value),
            # Loan balances (rows 11-13, 17-18, 22-23)
            "senior_loans": [],
            "subordinated_loans": [],
            "deposits": [],
        }
        # Senior loans
        for r in range(11, 14):
            dt = _date(ws.cell(row=r, column=2).value)
            principal = _dec(ws.cell(row=r, column=3).value)
            balance = _dec(ws.cell(row=r, column=4).value)
            if dt or principal or balance:
                result["financing"]["senior_loans"].append({
                    "date": dt, "principal": float(principal or 0), "balance": float(balance or 0),
                })
        # Subordinated
        for r in range(17, 19):
            dt = _date(ws.cell(row=r, column=2).value)
            principal = _dec(ws.cell(row=r, column=3).value)
            balance = _dec(ws.cell(row=r, column=4).value)
            if dt or principal or balance:
                result["financing"]["subordinated_loans"].append({
                    "date": dt, "principal": float(principal or 0), "balance": float(balance or 0),
                })
        # Deposits (PKM)
        for r in range(22, 24):
            dt = _date(ws.cell(row=r, column=2).value)
            principal = _dec(ws.cell(row=r, column=3).value)
            balance = _dec(ws.cell(row=r, column=4).value)
            if dt or principal or balance:
                result["financing"]["deposits"].append({
                    "date": dt, "principal": float(principal or 0), "balance": float(balance or 0),
                })
    else:
        warnings.append(f"טאב '{financing_tab}' לא נמצא")

    # ── 6. Programme (skip, store as reference) ──────────────

    if "פרוגרמה" in wb.sheetnames:
        result["programme"] = "found_but_skipped"

    # ── Summary ──────────────────────────────────────────────

    dev_res = result["apartments_developer_residential"]
    dev_com = result["apartments_developer_commercial"]
    res_res = result["apartments_resident_residential"]
    res_com = result["apartments_resident_commercial"]

    budget_lines = sum(len(items) for items in result["budget_categories"].values())
    budget_total = sum(
        float(item["cost_no_vat"] or 0)
        for items in result["budget_categories"].values()
        for item in items
    )

    guarantee_total = sum(g["indexed_balance"] for g in result["guarantees"])

    result["summary"] = {
        "developer_residential_count": len(dev_res),
        "developer_commercial_count": len(dev_com),
        "resident_residential_count": len(res_res),
        "resident_commercial_count": len(res_com),
        "total_apartments": len(dev_res) + len(dev_com) + len(res_res) + len(res_com),
        "budget_line_count": budget_lines,
        "budget_total": round(budget_total),
        "budget_categories_count": len(result["budget_categories"]),
        "guarantee_count": len(result["guarantees"]),
        "guarantee_total": round(guarantee_total),
        "has_financing": result["financing"] is not None,
    }

    return result
