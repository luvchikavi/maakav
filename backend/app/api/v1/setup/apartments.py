"""Apartment inventory setup endpoints."""

import io
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func

from ....database import get_db
from ....models.user import User
from ....models.project import Project
from ....models.apartment import Apartment, Ownership, UnitStatus, UnitType
from ....schemas.setup import ApartmentCreate, ApartmentResponse
from ....core.dependencies import get_current_user

import openpyxl
import pandas as pd

router = APIRouter(tags=["setup-apartments"])

OWNERSHIP_MAP = {"יזם": "developer", "דיירים": "resident", "בעלים": "resident"}
STATUS_MAP = {"לשיווק": "for_sale", "נמכר": "sold", "תמורה": "compensation", "להשכרה": "for_rent", "שמור": "reserved"}
TYPE_MAP = {
    "דירה": "apartment", "פנטהאוז": "penthouse", "גן": "garden", "דופלקס": "duplex",
    "משרדים": "office", "מסחר": "retail", "מחסן": "storage", "חניה": "parking",
}


@router.get("/projects/{project_id}/setup/apartments", response_model=list[ApartmentResponse])
async def list_apartments(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_project(project_id, user.firm_id, db)
    result = await db.execute(
        select(Apartment)
        .where(Apartment.project_id == project_id)
        .order_by(Apartment.building_number, Apartment.floor, Apartment.unit_number)
    )
    return result.scalars().all()


@router.post("/projects/{project_id}/setup/apartments", response_model=ApartmentResponse)
async def create_apartment(
    project_id: int,
    body: ApartmentCreate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_project(project_id, user.firm_id, db)
    apt = Apartment(project_id=project_id, **body.model_dump())

    # Auto-calc VAT if one price missing
    if apt.list_price_no_vat and not apt.list_price_with_vat:
        apt.list_price_with_vat = apt.list_price_no_vat * Decimal("1.18")
    elif apt.list_price_with_vat and not apt.list_price_no_vat:
        apt.list_price_no_vat = apt.list_price_with_vat / Decimal("1.18")

    # Resident units are always compensation
    if apt.ownership == Ownership.RESIDENT:
        apt.unit_status = UnitStatus.COMPENSATION
        apt.include_in_revenue = False

    db.add(apt)
    await db.commit()
    await db.refresh(apt)
    return apt


@router.delete("/projects/{project_id}/setup/apartments/{apartment_id}")
async def delete_apartment(
    project_id: int,
    apartment_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_project(project_id, user.firm_id, db)
    result = await db.execute(
        select(Apartment).where(Apartment.id == apartment_id, Apartment.project_id == project_id)
    )
    apt = result.scalar_one_or_none()
    if not apt:
        raise HTTPException(status_code=404, detail="הדירה לא נמצאה")
    await db.delete(apt)
    await db.commit()
    return {"ok": True}


@router.post("/projects/{project_id}/setup/apartments/upload")
async def upload_apartments_excel(
    project_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload Excel with apartment inventory. Auto-detects sheet and header row."""
    await _verify_project(project_id, user.firm_id, db)

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)

    results = {"imported": 0, "skipped": 0, "sheets_used": []}

    for sheet_name in wb.sheetnames:
        if not any(kw in sheet_name for kw in ["מלאי", "דירות", "מגורים", "מסחר"]):
            continue

        # Detect ownership from sheet name
        is_owner = "בעלים" in sheet_name
        default_ownership = "resident" if is_owner else "developer"

        header_row = _find_header_row(content, sheet_name)
        df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")

        results["sheets_used"].append(sheet_name)

        for _, row in df.iterrows():
            if row.notna().sum() < 3:
                continue

            apt_data = _parse_apartment_row(row, df.columns, default_ownership)
            if apt_data:
                apt = Apartment(project_id=project_id, **apt_data)
                db.add(apt)
                results["imported"] += 1
            else:
                results["skipped"] += 1

    wb.close()
    await db.commit()

    return results


def _parse_apartment_row(row, columns, default_ownership: str) -> dict | None:
    """Parse a single row into apartment data."""
    col_list = [str(c) for c in columns]

    def get(keywords):
        for kw in keywords:
            for col in col_list:
                if kw in col:
                    val = row.get(col)
                    if pd.notna(val):
                        return val
        return None

    floor = get(["קומה"])
    unit_num = get(["מס\"ד", "מספר"])
    if floor is None and unit_num is None:
        return None

    # Detect ownership
    ownership_val = get(["בעלות"])
    ownership = default_ownership
    if ownership_val:
        ownership = OWNERSHIP_MAP.get(str(ownership_val).strip(), default_ownership)

    # Detect status
    status_val = get(["תמורה", "לשיווק", "להשכרה"])
    status = "compensation" if ownership == "resident" else "for_sale"
    if status_val:
        status = STATUS_MAP.get(str(status_val).strip(), status)

    # Detect type
    type_val = get(["סוג נכס", "סוג"])
    unit_type = "apartment"
    if type_val:
        unit_type = TYPE_MAP.get(str(type_val).strip(), "apartment")

    return {
        "building_number": str(get(["בניין"]) or "A"),
        "floor": str(floor) if floor else None,
        "unit_number": str(unit_num) if unit_num else None,
        "unit_type": UnitType(unit_type),
        "ownership": Ownership(ownership),
        "unit_status": UnitStatus(status),
        "room_count": _to_decimal(get(["חדרים"])),
        "net_area_sqm": _to_decimal(get(["שטח פלדלת", "שטח נטו", "שטח"])),
        "balcony_area_sqm": _to_decimal(get(["מרפסת שמש"])),
        "terrace_area_sqm": _to_decimal(get(["מרפסת גג", "חצר"])),
        "parking_count": int(float(get(["חניה"]) or 0)),
        "storage_count": int(float(get(["מחסן"]) or 0)),
        "list_price_with_vat": _to_decimal(get(["שווי כולל מע"])),
        "list_price_no_vat": _to_decimal(get(["שווי ללא מע"])),
        "include_in_revenue": ownership == "developer",
    }


def _to_decimal(val) -> Decimal | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return Decimal(str(val))
    except Exception:
        return None


def _find_header_row(content: bytes, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name]
    markers = ["בניין", "קומה", "שטח", "סוג"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        text = " ".join(str(v).strip() if v else "" for v in row)
        if sum(1 for m in markers if m in text) >= 2:
            wb.close()
            return row_idx
    wb.close()
    return 0


async def _verify_project(project_id: int, firm_id: int, db: AsyncSession):
    result = await db.execute(
        select(Project).where(Project.id == project_id, Project.firm_id == firm_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="הפרויקט לא נמצא")
