"""Budget setup endpoints - Section 8 budget from Report 0."""

import io
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from sqlalchemy.orm import selectinload

from ....database import get_db
from ....models.user import User
from ....models.project import Project
from ....models.budget import BudgetCategory, BudgetLineItem, CategoryType
from ....schemas.setup import (
    BudgetCategoryCreate, BudgetCategoryResponse, BudgetUploadResponse,
)
from ....core.dependencies import get_current_user

import openpyxl
import pandas as pd

router = APIRouter(tags=["setup-budget"])

# Hebrew category name → CategoryType mapping
CATEGORY_MAPPING = {
    "קרקע והוצאות דיירים": CategoryType.TENANT_EXPENSES,
    "הוצאות דיירים": CategoryType.TENANT_EXPENSES,
    "קרקע ומיסוי": CategoryType.LAND_AND_TAXES,
    "כלליות": CategoryType.INDIRECT_COSTS,
    "הוצאות עקיפות": CategoryType.INDIRECT_COSTS,
    "הקמה": CategoryType.DIRECT_CONSTRUCTION,
    "בניה ישירה": CategoryType.DIRECT_CONSTRUCTION,
    "בניה ישירה - קבלן מבצע": CategoryType.DIRECT_CONSTRUCTION,
    "הוצאות חריגות": CategoryType.EXTRAORDINARY,
}

CATEGORY_ORDER = {
    CategoryType.TENANT_EXPENSES: 1,
    CategoryType.LAND_AND_TAXES: 2,
    CategoryType.INDIRECT_COSTS: 3,
    CategoryType.DIRECT_CONSTRUCTION: 4,
    CategoryType.EXTRAORDINARY: 5,
}


@router.get("/projects/{project_id}/setup/budget", response_model=list[BudgetCategoryResponse])
async def get_budget(
    project_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_project(project_id, user.firm_id, db)
    result = await db.execute(
        select(BudgetCategory)
        .options(selectinload(BudgetCategory.line_items))
        .where(BudgetCategory.project_id == project_id)
        .order_by(BudgetCategory.display_order)
    )
    return result.scalars().all()


@router.post("/projects/{project_id}/setup/budget", response_model=list[BudgetCategoryResponse])
async def save_budget(
    project_id: int,
    categories: list[BudgetCategoryCreate],
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Save or replace entire budget structure."""
    await _verify_project(project_id, user.firm_id, db)

    # Delete existing
    await db.execute(
        delete(BudgetLineItem).where(
            BudgetLineItem.category_id.in_(
                select(BudgetCategory.id).where(BudgetCategory.project_id == project_id)
            )
        )
    )
    await db.execute(delete(BudgetCategory).where(BudgetCategory.project_id == project_id))

    created = []
    for cat_data in categories:
        cat_type = CategoryType(cat_data.category_type)
        category = BudgetCategory(
            project_id=project_id,
            category_type=cat_type,
            display_order=CATEGORY_ORDER.get(cat_type, 99),
        )
        total = Decimal("0")
        for i, item_data in enumerate(cat_data.line_items):
            item = BudgetLineItem(
                line_number=i + 1,
                **item_data.model_dump(),
            )
            category.line_items.append(item)
            total += item.cost_no_vat
        category.total_amount = total
        db.add(category)
        created.append(category)

    await db.commit()

    # Reload with items
    result = await db.execute(
        select(BudgetCategory)
        .options(selectinload(BudgetCategory.line_items))
        .where(BudgetCategory.project_id == project_id)
        .order_by(BudgetCategory.display_order)
    )
    return result.scalars().all()


@router.post("/projects/{project_id}/setup/budget/upload", response_model=BudgetUploadResponse)
async def upload_budget_excel(
    project_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload Excel file with budget data (Section 8 format)."""
    await _verify_project(project_id, user.firm_id, db)

    content = await file.read()

    # Find the right sheet
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if any(kw in name for kw in ["תקציב", "הוצאות", "עלויות"]):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    wb.close()

    # Find header row
    header_row = _find_header_row(content, sheet_name)
    df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")

    # Map columns
    col_map = {}
    for col in df.columns:
        col_clean = str(col).strip()
        if "קטגוריה" in col_clean:
            col_map["category"] = col
        elif "סעיף" in col_clean or "תת קטגוריה" in col_clean:
            col_map["description"] = col
        elif "עלות ללא" in col_clean or "ללא מע" in col_clean:
            col_map["cost"] = col
        elif "שם ספק" in col_clean:
            col_map["supplier"] = col
        elif "הערה" in col_clean or "הערות" in col_clean:
            col_map["notes"] = col

    if "description" not in col_map and "cost" not in col_map:
        raise HTTPException(status_code=400, detail="לא זוהו עמודות תקציב בקובץ")

    # Parse rows into categories
    categories_dict: dict[str, list] = {}
    for _, row in df.iterrows():
        cat_name = str(row.get(col_map.get("category", ""), "")).strip() if "category" in col_map else ""
        desc = str(row.get(col_map.get("description", ""), "")).strip() if "description" in col_map else ""
        cost = row.get(col_map.get("cost", ""), 0) if "cost" in col_map else 0
        supplier = str(row.get(col_map.get("supplier", ""), "")).strip() if "supplier" in col_map else ""
        notes = str(row.get(col_map.get("notes", ""), "")).strip() if "notes" in col_map else ""

        if not desc or desc == "nan":
            continue
        if pd.isna(cost):
            cost = 0

        cat_name = cat_name if cat_name and cat_name != "nan" else "כלליות"

        if cat_name not in categories_dict:
            categories_dict[cat_name] = []
        categories_dict[cat_name].append({
            "description": desc,
            "cost_no_vat": float(cost),
            "supplier_name": supplier if supplier != "nan" else "",
            "notes": notes if notes != "nan" else "",
        })

    # Delete existing and create new
    await db.execute(
        delete(BudgetLineItem).where(
            BudgetLineItem.category_id.in_(
                select(BudgetCategory.id).where(BudgetCategory.project_id == project_id)
            )
        )
    )
    await db.execute(delete(BudgetCategory).where(BudgetCategory.project_id == project_id))

    total_items = 0
    total_budget = 0.0
    for cat_name, items in categories_dict.items():
        cat_type = CATEGORY_MAPPING.get(cat_name, CategoryType.INDIRECT_COSTS)
        category = BudgetCategory(
            project_id=project_id,
            category_type=cat_type,
            display_order=CATEGORY_ORDER.get(cat_type, 99),
        )
        cat_total = Decimal("0")
        for i, item in enumerate(items):
            line = BudgetLineItem(
                line_number=i + 1,
                description=item["description"],
                cost_no_vat=Decimal(str(item["cost_no_vat"])),
                supplier_name=item["supplier_name"],
                notes=item["notes"],
            )
            category.line_items.append(line)
            cat_total += line.cost_no_vat
            total_items += 1
        category.total_amount = cat_total
        total_budget += float(cat_total)
        db.add(category)

    await db.commit()

    return BudgetUploadResponse(
        categories_count=len(categories_dict),
        items_count=total_items,
        total_budget=total_budget,
    )


def _find_header_row(content: bytes, sheet_name: str) -> int:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name]
    markers = ["קטגוריה", "סעיף", "עלות"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        text = " ".join(str(v).strip() if v else "" for v in row).lower()
        if sum(1 for m in markers if m in text) >= 2:
            wb.close()
            return row_idx
    wb.close()
    return 0


async def _verify_project(project_id: int, firm_id: int, db: AsyncSession) -> Project:
    result = await db.execute(
        select(Project).where(Project.id == project_id, Project.firm_id == firm_id)
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="הפרויקט לא נמצא")
    return project
