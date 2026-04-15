"""Check approval endpoints — upload monthly tab, CRUD, auto-match, budget alerts."""

from datetime import date, datetime
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from pydantic import BaseModel

from ....database import get_db
from ....models.user import User
from ....models.project import Project
from ....models.monthly_report import MonthlyReport
from ....models.payment_approval import PaymentApproval, OperationType, ApprovalStatus, PaymentApprovalStatus
from ....models.bank_statement import BankTransaction, TransactionType
from ....models.budget import BudgetCategory, BudgetLineItem
from ....models.budget_tracking import BudgetTrackingSnapshot, BudgetTrackingLine
from ....core.dependencies import get_current_user

router = APIRouter(tags=["check-approvals"])


# ── List / Summary ───────────────────────────────────────────


@router.get("/projects/{project_id}/monthly-reports/{report_id}/checks")
async def list_checks(
    project_id: int, report_id: int,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """List all check approvals for a monthly report."""
    report = await _get_report(project_id, report_id, user.firm_id, db)

    checks = (await db.execute(
        select(PaymentApproval)
        .where(PaymentApproval.project_id == project_id, PaymentApproval.monthly_report_id == report_id)
        .order_by(PaymentApproval.serial_number)
    )).scalars().all()

    return [_serialize(c) for c in checks]


@router.get("/projects/{project_id}/monthly-reports/{report_id}/checks/summary")
async def checks_summary(
    project_id: int, report_id: int,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """Summary: totals, counts, budget alerts."""
    report = await _get_report(project_id, report_id, user.firm_id, db)

    checks = (await db.execute(
        select(PaymentApproval)
        .where(PaymentApproval.project_id == project_id, PaymentApproval.monthly_report_id == report_id)
    )).scalars().all()

    total_approved = sum(float(c.amount_with_vat) for c in checks if c.approval_status == ApprovalStatus.APPROVED)
    total_pending = sum(float(c.amount_with_vat) for c in checks if c.approval_status == ApprovalStatus.PENDING)
    total_paid = sum(float(c.amount_with_vat) for c in checks if c.payment_status == PaymentApprovalStatus.PAID)

    # Budget alerts — check each category usage
    alerts = await _get_budget_alerts(project_id, report_id, checks, db)

    return {
        "total_checks": len(checks),
        "pending_count": sum(1 for c in checks if c.approval_status == ApprovalStatus.PENDING),
        "approved_count": sum(1 for c in checks if c.approval_status == ApprovalStatus.APPROVED),
        "paid_count": sum(1 for c in checks if c.payment_status == PaymentApprovalStatus.PAID),
        "total_approved_amount": round(total_approved),
        "total_pending_amount": round(total_pending),
        "total_paid_amount": round(total_paid),
        "budget_alerts": alerts,
    }


# ── Upload from Excel tab ────────────────────────────────────


@router.post("/projects/{project_id}/monthly-reports/{report_id}/checks/upload")
async def upload_checks(
    project_id: int, report_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """Upload a monthly check approval Excel tab. Parses and creates records."""
    report = await _get_report(project_id, report_id, user.firm_id, db)
    content = await file.read()

    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find the "שיקים לתשלום בחודש" section
    data_start = None
    for r in range(1, min(ws.max_row + 1, 20)):
        val = str(ws.cell(row=r, column=7).value or "")
        if "שיקים לתשלום" in val or "מס' סידורי" in val:
            data_start = r + 1
            break

    if data_start is None:
        # Try header detection
        for r in range(1, min(ws.max_row + 1, 20)):
            val = str(ws.cell(row=r, column=7).value or "")
            if "סידורי" in val:
                data_start = r + 1
                break

    if data_start is None:
        data_start = 14  # Default based on format

    created = 0
    for r in range(data_start, ws.max_row + 1):
        serial = ws.cell(row=r, column=7).value
        if serial is None or not str(serial).strip().isdigit():
            continue

        op_type_raw = str(ws.cell(row=r, column=8).value or "").strip()
        op_type = OperationType.CHECK if "שיק" in op_type_raw else OperationType.TRANSFER

        amount_raw = ws.cell(row=r, column=9).value
        if amount_raw is None:
            continue
        try:
            amount = Decimal(str(amount_raw).replace(",", ""))
        except Exception:
            continue

        beneficiary = str(ws.cell(row=r, column=10).value or "").strip()
        if not beneficiary:
            continue

        due_date_raw = ws.cell(row=r, column=11).value
        due_dt = None
        if isinstance(due_date_raw, datetime):
            due_dt = due_date_raw.date()
        elif isinstance(due_date_raw, date):
            due_dt = due_date_raw

        description = str(ws.cell(row=r, column=12).value or "").strip()
        budget_cat = str(ws.cell(row=r, column=13).value or "").strip()
        vat_rate_raw = ws.cell(row=r, column=15).value
        amount_no_vat_raw = ws.cell(row=r, column=16).value
        vat_amount_raw = ws.cell(row=r, column=17).value
        invoice = str(ws.cell(row=r, column=19).value or "").strip()
        reference = str(ws.cell(row=r, column=20).value or "").strip()

        check = PaymentApproval(
            project_id=project_id,
            monthly_report_id=report_id,
            report_month=report.report_month,
            operation_type=op_type,
            serial_number=int(serial),
            amount_with_vat=amount,
            amount_no_vat=Decimal(str(amount_no_vat_raw).replace(",", "")) if amount_no_vat_raw else None,
            vat_amount=Decimal(str(vat_amount_raw).replace(",", "")) if vat_amount_raw else None,
            beneficiary_name=beneficiary,
            due_date=due_dt,
            description=description or None,
            budget_category=budget_cat or None,
            invoice_number=invoice or None,
            reference=reference or None,
            approval_status=ApprovalStatus.PENDING,
            created_by=user.id,
        )
        db.add(check)
        created += 1

    await db.commit()

    # Auto-match with bank transactions
    matched = await _auto_match(project_id, report_id, db)

    return {"created": created, "auto_matched": matched}


# ── CRUD ─────────────────────────────────────────────────────


class CheckApproveRequest(BaseModel):
    approval_status: str  # "approved" or "rejected"
    notes: str | None = None


@router.post("/projects/{project_id}/monthly-reports/{report_id}/checks/{check_id}/approve")
async def approve_check(
    project_id: int, report_id: int, check_id: int,
    body: CheckApproveRequest,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """Approve or reject a check."""
    check = await _get_check(check_id, project_id, db)
    check.approval_status = ApprovalStatus(body.approval_status)
    check.approved_by = f"{user.first_name} {user.last_name}"
    check.approved_date = datetime.utcnow()
    if body.notes:
        check.approval_notes = body.notes
    await db.commit()
    return {"ok": True}


@router.delete("/projects/{project_id}/monthly-reports/{report_id}/checks/{check_id}")
async def delete_check(
    project_id: int, report_id: int, check_id: int,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    check = await _get_check(check_id, project_id, db)
    await db.delete(check)
    await db.commit()
    return {"ok": True}


# ── Auto-generate drafts from bank transactions ──────────────


@router.post("/projects/{project_id}/monthly-reports/{report_id}/checks/auto-generate")
async def auto_generate_checks(
    project_id: int, report_id: int,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """
    Auto-generate draft check records from bank debit transactions.
    Creates a PaymentApproval for each DEBIT transaction that doesn't already have a match.
    Status = APPROVED + PAID (since it already appeared in bank).
    """
    report = await _get_report(project_id, report_id, user.firm_id, db)

    # Get all debit transactions for this report
    debit_txs = (await db.execute(
        select(BankTransaction).where(
            BankTransaction.monthly_report_id == report_id,
            BankTransaction.project_id == project_id,
            BankTransaction.transaction_type == TransactionType.DEBIT,
        ).order_by(BankTransaction.transaction_date)
    )).scalars().all()

    if not debit_txs:
        return {"generated": 0, "skipped": 0}

    # Get existing checks to avoid duplicates
    existing = (await db.execute(
        select(PaymentApproval).where(
            PaymentApproval.project_id == project_id,
            PaymentApproval.monthly_report_id == report_id,
        )
    )).scalars().all()

    # Build set of already-linked transaction IDs
    linked_tx_ids = {c.linked_transaction_id for c in existing if c.linked_transaction_id}
    # Also match by amount+date to avoid near-duplicates
    existing_keys = {
        (float(c.amount_with_vat), str(c.due_date)) for c in existing
    }

    generated = 0
    skipped = 0
    serial = len(existing) + 1

    for tx in debit_txs:
        # Skip if already linked
        if tx.id in linked_tx_ids:
            skipped += 1
            continue

        # Skip if same amount+date already exists
        key = (float(tx.amount), str(tx.transaction_date))
        if key in existing_keys:
            skipped += 1
            continue

        # Map category to budget category name
        budget_cat = _tx_category_to_budget(tx.category.value if tx.category else None)

        check = PaymentApproval(
            project_id=project_id,
            monthly_report_id=report_id,
            report_month=report.report_month,
            operation_type=OperationType.TRANSFER,
            serial_number=serial,
            amount_with_vat=tx.amount,
            amount_no_vat=Decimal(str(round(float(tx.amount) / 1.18, 2))),
            beneficiary_name=tx.description or "לא ידוע",
            due_date=tx.transaction_date,
            description=tx.description,
            budget_category=budget_cat,
            reference=tx.reference_number,
            approval_status=ApprovalStatus.APPROVED,
            payment_status=PaymentApprovalStatus.PAID,
            linked_transaction_id=tx.id,
            actual_payment_date=tx.transaction_date,
            created_by=user.id,
        )
        db.add(check)
        linked_tx_ids.add(tx.id)
        existing_keys.add(key)
        serial += 1
        generated += 1

    if generated:
        await db.commit()

    return {"generated": generated, "skipped": skipped}


def _tx_category_to_budget(cat: str | None) -> str | None:
    """Map transaction category to Hebrew budget category name."""
    mapping = {
        "direct_construction": "בניה ישירה",
        "indirect_costs": "הוצאות עקיפות",
        "tenant_expenses": "הוצאות דיירים",
        "land_and_taxes": "קרקע ומיסוי",
        "interest_and_fees": "ריביות ועמלות",
        "loan_repayment": "החזר הלוואה",
        "deposit_to_savings": "הפקדה לפקדון",
        "other_expense": "הוצאה אחרת",
    }
    return mapping.get(cat) if cat else None


# ── Expense Forecast ─────────────────────────────────────────


@router.get("/projects/{project_id}/expense-forecast")
async def get_expense_forecast(
    project_id: int,
    user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """
    Expense forecast: unpaid approved checks + future scheduled payments.
    Shows what's expected to be paid in coming months.
    """
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.firm_id == user.firm_id)
    )).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="הפרויקט לא נמצא")

    # 1. Unpaid approved checks (across all reports)
    unpaid_checks = (await db.execute(
        select(PaymentApproval).where(
            PaymentApproval.project_id == project_id,
            PaymentApproval.approval_status == ApprovalStatus.APPROVED,
            PaymentApproval.payment_status == PaymentApprovalStatus.UNPAID,
        ).order_by(PaymentApproval.due_date)
    )).scalars().all()

    # 2. Pending checks (not yet approved)
    pending_checks = (await db.execute(
        select(PaymentApproval).where(
            PaymentApproval.project_id == project_id,
            PaymentApproval.approval_status == ApprovalStatus.PENDING,
        ).order_by(PaymentApproval.due_date)
    )).scalars().all()

    # 3. Future scheduled payments from sales (buyer installments)
    from ....models.sales import PaymentScheduleItem, PaymentStatus, SalesContract
    future_payments = (await db.execute(
        select(PaymentScheduleItem, SalesContract.buyer_name)
        .join(SalesContract, SalesContract.id == PaymentScheduleItem.contract_id)
        .where(
            SalesContract.project_id == project_id,
            PaymentScheduleItem.status.in_([PaymentStatus.SCHEDULED, PaymentStatus.PARTIAL]),
        )
        .order_by(PaymentScheduleItem.scheduled_date)
    )).all()

    # 4. Budget status per category
    from ....models.budget_tracking import BudgetTrackingSnapshot, BudgetTrackingLine
    # Get latest snapshot
    latest_snapshot = (await db.execute(
        select(BudgetTrackingSnapshot)
        .where(BudgetTrackingSnapshot.project_id == project_id)
        .order_by(BudgetTrackingSnapshot.id.desc())
        .limit(1)
    )).scalar_one_or_none()

    budget_status = []
    if latest_snapshot:
        lines = (await db.execute(
            select(BudgetTrackingLine).where(BudgetTrackingLine.snapshot_id == latest_snapshot.id)
        )).scalars().all()
        for line in lines:
            pct = float(line.execution_percent) if line.execution_percent else 0
            remaining = float(line.remaining_indexed) if line.remaining_indexed else 0
            budget_status.append({
                "category": line.category,
                "total_budget": float(line.total_indexed) if line.total_indexed else 0,
                "cumulative_paid": float(line.cumulative_actual) if line.cumulative_actual else 0,
                "remaining": remaining,
                "usage_percent": round(pct, 1),
                "alert": "error" if pct >= 100 else "warning" if pct >= 90 else None,
            })

    return {
        "unpaid_checks": [_serialize(c) for c in unpaid_checks],
        "pending_checks": [_serialize(c) for c in pending_checks],
        "future_receipts": [
            {
                "buyer_name": row.buyer_name,
                "amount": float(row.PaymentScheduleItem.scheduled_amount),
                "date": str(row.PaymentScheduleItem.scheduled_date),
                "description": row.PaymentScheduleItem.description,
            }
            for row in future_payments
        ],
        "budget_status": budget_status,
        "summary": {
            "total_unpaid": round(sum(float(c.amount_with_vat) for c in unpaid_checks)),
            "total_pending": round(sum(float(c.amount_with_vat) for c in pending_checks)),
            "total_future_receipts": round(sum(float(row.PaymentScheduleItem.scheduled_amount) for row in future_payments)),
            "unpaid_count": len(unpaid_checks),
            "pending_count": len(pending_checks),
            "future_receipts_count": len(future_payments),
        },
    }


# ── Auto-match with bank transactions ────────────────────────


async def _auto_match(project_id: int, report_id: int, db: AsyncSession) -> int:
    """
    Match checks with bank transactions using multi-criteria scoring:
    - Check number / reference (40 points)
    - Amount match (40 points)
    - Beneficiary name (20 points)
    - Date proximity (10 points)
    Minimum score: 60 to auto-match
    """
    unmatched = (await db.execute(
        select(PaymentApproval).where(
            PaymentApproval.project_id == project_id,
            PaymentApproval.monthly_report_id == report_id,
            PaymentApproval.payment_status == PaymentApprovalStatus.UNPAID,
        )
    )).scalars().all()

    if not unmatched:
        return 0

    # Get debit transactions for this report
    bank_txs = (await db.execute(
        select(BankTransaction).where(
            BankTransaction.monthly_report_id == report_id,
            BankTransaction.transaction_type == TransactionType.DEBIT,
        )
    )).scalars().all()

    # Also check previous reports for carryover checks
    prev_txs = (await db.execute(
        select(BankTransaction).where(
            BankTransaction.project_id == project_id,
            BankTransaction.transaction_type == TransactionType.DEBIT,
        )
    )).scalars().all()

    all_txs = list({tx.id: tx for tx in list(bank_txs) + list(prev_txs)}.values())

    matched = 0
    used_tx_ids = set()

    # Already linked
    existing_links = (await db.execute(
        select(PaymentApproval.linked_transaction_id).where(
            PaymentApproval.project_id == project_id,
            PaymentApproval.linked_transaction_id.is_not(None),
        )
    )).scalars().all()
    used_tx_ids = set(existing_links)

    for check in unmatched:
        best_tx = None
        best_score = 0

        for tx in all_txs:
            if tx.id in used_tx_ids:
                continue

            score = 0

            # 1. Check number match (40 points)
            if check.check_number and tx.reference_number:
                if check.check_number.strip() == tx.reference_number.strip():
                    score += 40
                elif check.check_number.strip() in tx.reference_number or tx.reference_number.strip() in str(check.check_number):
                    score += 25

            # 2. Amount match (40 points)
            amount_diff = abs(float(check.amount_with_vat) - float(tx.amount))
            if amount_diff < 1:
                score += 40
            elif amount_diff < float(check.amount_with_vat) * 0.02:  # 2% tolerance
                score += 25
            elif amount_diff < float(check.amount_with_vat) * 0.05:  # 5% tolerance
                score += 10

            # 3. Beneficiary name match (20 points)
            if check.beneficiary_name and tx.description:
                check_name = check.beneficiary_name.strip().lower()
                tx_desc = tx.description.strip().lower()
                if check_name in tx_desc or tx_desc in check_name:
                    score += 20
                elif any(word in tx_desc for word in check_name.split() if len(word) > 2):
                    score += 10

            # 4. Date proximity (10 points)
            if check.due_date and tx.transaction_date:
                from datetime import timedelta
                days_diff = abs((check.due_date - tx.transaction_date).days)
                if days_diff <= 1:
                    score += 10
                elif days_diff <= 7:
                    score += 5
                elif days_diff <= 14:
                    score += 2

            if score > best_score:
                best_score = score
                best_tx = tx

        # Apply match if score >= 60
        if best_tx and best_score >= 60:
            check.linked_transaction_id = best_tx.id
            check.payment_status = PaymentApprovalStatus.PAID
            check.actual_payment_date = best_tx.transaction_date
            check.approval_status = ApprovalStatus.APPROVED
            used_tx_ids.add(best_tx.id)
            matched += 1

    if matched:
        await db.commit()
    return matched


# ── Budget alerts ────────────────────────────────────────────

CATEGORY_MAP = {
    "בניה ישירה": "direct_construction",
    "קבלן מבצע": "direct_construction",
    "הוצאות עקיפות": "indirect_costs",
    "כלליות": "indirect_costs",
    "קרקע": "land_and_taxes",
    "דיירים": "tenant_expenses",
    "הובלה": "tenant_expenses",
    "חיבור חשמל": "indirect_costs",
    "תכנון": "indirect_costs",
    "שיווק": "indirect_costs",
    "משפטי": "indirect_costs",
    "פיקוח": "indirect_costs",
    "ביטוח": "indirect_costs",
    "מימון": "indirect_costs",
}


async def _get_budget_alerts(project_id: int, report_id: int, checks: list, db: AsyncSession) -> list:
    """Check if any budget category is over or near its limit."""
    alerts = []

    # Get latest budget tracking
    snapshot = (await db.execute(
        select(BudgetTrackingSnapshot).where(BudgetTrackingSnapshot.monthly_report_id == report_id)
    )).scalar_one_or_none()

    if not snapshot:
        return alerts

    lines = (await db.execute(
        select(BudgetTrackingLine).where(BudgetTrackingLine.snapshot_id == snapshot.id)
    )).scalars().all()

    for line in lines:
        pct = float(line.execution_percent) if line.execution_percent else 0
        remaining = float(line.remaining_indexed) if line.remaining_indexed else 0

        # Check approved checks that will reduce this category further
        pending_for_cat = sum(
            float(c.amount_with_vat)
            for c in checks
            if c.approval_status in (ApprovalStatus.PENDING, ApprovalStatus.APPROVED)
            and c.payment_status == PaymentApprovalStatus.UNPAID
            and _matches_category(c.budget_category, line.category)
        )

        effective_remaining = remaining - pending_for_cat
        effective_pct = pct + (pending_for_cat / float(line.total_indexed) * 100 if line.total_indexed and float(line.total_indexed) > 0 else 0)

        if effective_pct >= 100:
            alerts.append({
                "category": line.category,
                "severity": "error",
                "message": f"חריגת תקציב בסעיף {line.category}: ניצול {effective_pct:.0f}%",
                "usage_percent": round(effective_pct, 1),
                "remaining": round(effective_remaining),
            })
        elif effective_pct >= 90:
            alerts.append({
                "category": line.category,
                "severity": "warning",
                "message": f"קרוב למימוש בסעיף {line.category}: ניצול {effective_pct:.0f}%",
                "usage_percent": round(effective_pct, 1),
                "remaining": round(effective_remaining),
            })

    return alerts


def _matches_category(check_cat: str | None, budget_cat: str) -> bool:
    if not check_cat:
        return False
    check_lower = check_cat.lower().strip()
    # Direct match
    if budget_cat in check_lower:
        return True
    # Map
    for keyword, mapped in CATEGORY_MAP.items():
        if keyword in check_lower and mapped == budget_cat:
            return True
    return False


# ── Helpers ──────────────────────────────────────────────────


def _serialize(c: PaymentApproval) -> dict:
    return {
        "id": c.id,
        "serial_number": c.serial_number,
        "operation_type": c.operation_type.value,
        "check_number": c.check_number,
        "amount_with_vat": float(c.amount_with_vat),
        "amount_no_vat": float(c.amount_no_vat) if c.amount_no_vat else None,
        "vat_amount": float(c.vat_amount) if c.vat_amount else None,
        "beneficiary_name": c.beneficiary_name,
        "due_date": str(c.due_date) if c.due_date else None,
        "description": c.description,
        "budget_category": c.budget_category,
        "invoice_number": c.invoice_number,
        "reference": c.reference,
        "approval_status": c.approval_status.value,
        "payment_status": c.payment_status.value,
        "approved_by": c.approved_by,
        "approved_date": str(c.approved_date) if c.approved_date else None,
    }


async def _get_report(project_id: int, report_id: int, firm_id: int, db: AsyncSession) -> MonthlyReport:
    project = (await db.execute(
        select(Project).where(Project.id == project_id, Project.firm_id == firm_id)
    )).scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="הפרויקט לא נמצא")
    report = (await db.execute(
        select(MonthlyReport).where(MonthlyReport.id == report_id, MonthlyReport.project_id == project_id)
    )).scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="הדוח לא נמצא")
    return report


async def _get_check(check_id: int, project_id: int, db: AsyncSession) -> PaymentApproval:
    check = (await db.execute(
        select(PaymentApproval).where(PaymentApproval.id == check_id, PaymentApproval.project_id == project_id)
    )).scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="השיק לא נמצא")
    return check
